from openpyxl import load_workbook

wb = load_workbook('ventas.xlsx')
productos = wb['Productos']
tiendas = wb['Tiendas']

SKU = input("Ingresa el codigo SKU: ")
for row in productos.iter_rows(min_row=2, values_only=True):
    if row[0] == SKU:
        i = 0
        while i <= 2:
            print(row[i])
            i += 1
        break
else:
    print("Producto no registrado")

from openpyxl import load_workbook

wb = load_workbook('ventas.xlsx')

tienda = wb['Tiendas']

tiendaEliminada = input("Ingresa el SKU del prodcuto que deseas eliminar ")
# Eliminar el registro con índice 2
for row in tienda.iter_rows():
    for cell in row:
        if cell.value == tiendaEliminada:
            tienda.delete_rows(cell.row)


# Guardar los cambios en un nuevo archivo xlsx
wb.save('ventas.xlsx')

# Mostrar la tiendas actualizada
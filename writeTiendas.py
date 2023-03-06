from openpyxl import load_workbook

filesheet = 'ventas.xlsx'
wb = load_workbook(filesheet)

Tiendas = wb.active

num_filas = Tiendas.max_row

SKU = input("Ingresa el codigo SKU ")
Nombre = input("Ingresa el nombre del producto ")
Unidad = input("Ingresa el valor de Unidad ")

Tiendas.cell(row=num_filas+1, column=1, value=SKU)
Tiendas.cell(row=num_filas+1, column=2, value=Nombre)
Tiendas.cell(row=num_filas+1, column=3, value=Unidad)

wb.save('ventas.xlsx')
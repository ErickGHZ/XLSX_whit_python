import openpyxl
from openpyxl import Workbook

workbook = Workbook()
  

workbook.save(filename="demo.xlsx")

wb = openpyxl.Workbook()
productos = wb.active
a1 = productos.cell(row = 1, column = 1)
a2 = productos.cell(row = 1, column = 2)
a3 = productos.cell(row = 1, column = 3)
b1 = productos.cell(row = 2, column = 1)
b2 = productos.cell(row = 2, column = 2)
b3 = productos.cell(row = 2, column = 3)
c1 = productos.cell(row = 3, column = 1)
c2 = productos.cell(row = 3, column = 2)
c3 = productos.cell(row = 3, column = 3)

a1.value = "SKU"
a2.value = "1"
a3.value = "2"
b1.value = "Nombre"
b2.value = "Producto1"
b3.value = "Producto2"
c1.value = "Unidad"
c2.value = "Pieza"
c3.value = "Pieza"
wb.save("demo.xlsx") 


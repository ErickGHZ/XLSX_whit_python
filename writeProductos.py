from openpyxl import load_workbook

filesheet = 'ventas.xlsx'
wb = load_workbook(filesheet)

Productos = wb.active

num_filas = Productos.max_row

SKU = input("Ingresa el codigo SKU ")
Nombre = input("Ingresa el nombre del producto ")
Unidad = input("Ingresa el valor de Unidad ")

Productos.cell(row=num_filas+1, column=1, value=SKU)
Productos.cell(row=num_filas+1, column=2, value=Nombre)
Productos.cell(row=num_filas+1, column=3, value=Unidad)

wb.save('ventas.xlsx')
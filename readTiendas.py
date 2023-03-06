from openpyxl import load_workbook

wb = load_workbook('ventas.xlsx')

tiendas = wb['Tiendas']

ID = input("Ingresa el codigo ID: ")
for row in tiendas.iter_rows(min_row=2, values_only=True):
    if row[0] == ID:
        i = 0
        while i <= 5:
            print(row[i])
            i += 1
        break
else:
    print("Tienda no registrada")

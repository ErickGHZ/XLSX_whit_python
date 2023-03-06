from openpyxl import load_workbook

wb = load_workbook('ventas.xlsx')

productos = wb['Productos']

productoEliminado = input("Ingresa el SKU del prodcuto que deseas eliminar ")
encontrado = False
for row in productos.iter_rows():
    for cell in row:
        if cell.value == productoEliminado:
            productos.delete_rows(cell.row)
            print("Producto eliminado correctamente")
            encontrado = True
            break

if encontrado is False:  #  else en pocas palabras
    print("Producto no encontrado")
        


# Guardar los cambios en un nuevo archivo xlsx
wb.save('ventas.xlsx')

 